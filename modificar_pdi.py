import os
import shutil
import openpyxl
import sys
import subprocess
import zipfile
import re
import tempfile

def check_and_elevate():
    """Eleva los privilegios del script a Administrador en Windows si es necesario."""
    if os.name == 'nt':
        import ctypes
        try:
            if not ctypes.windll.shell32.IsUserAnAdmin():
                log_message("Elevando privilegios a Administrador para configurar la impresora a doble cara...")
                params = " ".join([f'"{arg}"' for arg in sys.argv[1:]])
                if getattr(sys, 'frozen', False):
                    # Si está compilado con PyInstaller (.exe)
                    ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, params, None, 1)
                else:
                    # Si se ejecuta como script .py
                    script_path = os.path.abspath(sys.argv[0])
                    ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, f'"{script_path}" {params}', None, 1)
                sys.exit(0)
        except Exception as e:
            log_message(f"Error al intentar elevar privilegios automáticamente: {e}")
            log_message("Por favor, ejecuta este programa haciendo clic derecho y seleccionando 'Ejecutar como administrador'.")
            input("\nPresiona Enter para continuar...")

def log_message(msg):
    """Guarda un mensaje en el archivo de depuración local sin mostrarlo en la consola."""
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        log_file = os.path.join(script_dir, "pdi_debug_log.txt")
        with open(log_file, "a", encoding="utf-8") as f:
            f.write(msg + "\n")
    except Exception:
        pass

def log_user(msg):
    """Muestra un mensaje en la consola y lo guarda en el archivo de depuración."""
    print(msg)
    log_message(msg)

def strip_printer_settings_from_zip(xlsx_path):
    """Desempaqueta el archivo Excel (.xlsx), elimina las configuraciones de impresora incrustadas (.bin) y limpia sus relaciones XML."""
    log_message("Limpiando la caché de configuración de impresora del archivo Excel...")
    temp_dir = tempfile.mkdtemp()
    temp_zip_path = os.path.join(temp_dir, "temp.zip")
    
    try:
        shutil.copyfile(xlsx_path, temp_zip_path)
        
        with zipfile.ZipFile(temp_zip_path, 'r') as zin:
            with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    filename = item.filename
                    
                    if "printerSettings" in filename:
                        continue
                        
                    data = zin.read(filename)
                    
                    if filename.startswith("xl/worksheets/_rels/") and filename.endswith(".xml.rels"):
                        data_str = data.decode('utf-8', errors='ignore')
                        data_str = re.sub(
                            r'<Relationship[^>]*relationships/printerSettings[^>]*/>', 
                            '', 
                            data_str
                        )
                        data = data_str.encode('utf-8')
                        
                    elif filename.startswith("xl/worksheets/sheet") and filename.endswith(".xml"):
                        data_str = data.decode('utf-8', errors='ignore')
                        
                        def clean_pagesetup(match):
                            tag = match.group(0)
                            tag_cleaned = re.sub(r'\s+r:id="[^"]+"', '', tag)
                            return tag_cleaned
                        
                        data_str = re.sub(r'<pageSetup[^>]+/>', clean_pagesetup, data_str)
                        data = data_str.encode('utf-8')
                        
                    zout.writestr(item, data)
        log_message("--> Caché de impresora limpiada exitosamente del Excel.")
    except Exception as e:
        log_message(f"--> Advertencia al limpiar caché del Excel: {e}")
    finally:
        try:
            shutil.rmtree(temp_dir)
        except Exception:
            pass

def is_jimny_sheet_match(wb, sheet_name, jimny_doors):
    if not jimny_doors:
        return True
    try:
        sheet = wb[sheet_name]
        d9_val = str(sheet['D9'].value or '').upper()
        is_5_door = ('5' in d9_val or '5PTAS' in d9_val or '5 PTS' in d9_val)
        if jimny_doors == "5" and is_5_door:
            return True
        if jimny_doors == "3" and not is_5_door:
            return True
    except Exception:
        pass
    return False

def get_excel_printer_string(printer_name):
    """Encuentra el puerto (ej. 'on Ne01:') en el registro de Windows para construir la cadena ActivePrinter de Excel."""
    if os.name != 'nt':
        return printer_name
    try:
        import winreg
        for key_path in [
            r"Software\Microsoft\Windows NT\CurrentVersion\Devices",
            r"Software\Microsoft\Windows NT\CurrentVersion\PrinterPorts"
        ]:
            try:
                key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, key_path)
                i = 0
                while True:
                    try:
                        name, value, _ = winreg.EnumValue(key, i)
                        if name.lower().strip() == printer_name.lower().strip():
                            parts = value.split(",")
                            if len(parts) >= 2:
                                port = parts[1]
                                # Asegurar que el puerto termine con dos puntos (esencial para Excel ActivePrinter)
                                if not port.endswith(":"):
                                    port += ":"
                                winreg.CloseKey(key)
                                return f"{name} on {port}"
                        i += 1
                    except OSError:
                        break
                winreg.CloseKey(key)
            except Exception:
                continue
    except Exception as e:
        log_message(f"Advertencia al leer el puerto de la impresora del registro: {e}")
        
    return printer_name

def set_excel_active_printer(excel_app, printer_name):
    """Detecta dinámicamente el idioma de Excel (on / en) y configura la propiedad ActivePrinter con su puerto."""
    # 1. Intentar configurar la impresora directamente
    try:
        excel_app.ActivePrinter = printer_name
        log_message(f"Impresora activa configurada directamente en Excel: '{printer_name}'")
        return True
    except Exception:
        pass

    # 2. Detectar dinámicamente si Excel espera " on " (Inglés) o " en " (Español)
    separator = " on "
    try:
        current_active = excel_app.ActivePrinter
        log_message(f"ActivePrinter actual detectado en Excel: '{current_active}'")
        if " en " in current_active:
            separator = " en "
            log_message("Excel en Español detectado de manera dinámica.")
        else:
            log_message("Excel en Inglés detectado de manera dinámica.")
    except Exception as e:
        log_message(f"No se pudo consultar el ActivePrinter actual: {e}")

    # 3. Intentar construir la cadena usando el puerto real leído del registro de Windows
    excel_printer_str = get_excel_printer_string(printer_name)
    if excel_printer_str != printer_name:
        # Reemplazar el separador si detectamos español
        final_printer_str = excel_printer_str
        if separator == " en " and " on " in excel_printer_str:
            final_printer_str = excel_printer_str.replace(" on ", " en ")
        
        try:
            excel_app.ActivePrinter = final_printer_str
            log_message(f"Impresora activa configurada desde Puerto de Registro: '{final_printer_str}'")
            return True
        except Exception as e:
            log_message(f"Fallo al asignar la impresora del registro '{final_printer_str}': {e}")

    # 4. Fallback de barrido de puertos NeXX como último recurso
    for i in range(100):
        try:
            port_str = f"{printer_name}{separator}Ne{i:02d}:"
            excel_app.ActivePrinter = port_str
            log_message(f"Impresora activa configurada mediante barrido NeXX: '{port_str}'")
            return True
        except Exception:
            pass

    raise Exception("No se pudo asociar la impresora con ningún puerto de Excel.")

def ensure_duplex_printer_profile():
    """Detecta o crea una copia de la impresora predeterminada configurada a doble cara en el sistema."""
    if os.name != 'nt':
        return None
    try:
        import win32print
        import win32con
        
        # 1. Buscar si ya existe una impresora que contenga DOBLE CARA o DUPLEX
        printers = win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS)
        for _, name, _, _ in printers:
            name_up = name.upper()
            if "DOBLE CARA" in name_up or "DUPLEX" in name_up:
                log_message(f"--> Impresora especial de doble cara preconfigurada detectada: '{name}'")
                return name
                
        # 2. Si no existe, preguntar si se desea crear una copia configurada a doble cara
        default_printer = win32print.GetDefaultPrinter()
        print("\n" + "=" * 60)
        print("CONFIGURACIÓN AUTOMÁTICA DE DOBLE CARA")
        print("=" * 60)
        print(f"Para asegurar la impresión a doble cara en tu impresora Kyocera,")
        print(f"podemos crear una copia de tu impresora '{default_printer}'")
        print("configurada permanentemente a doble cara en Windows.")
        print("-" * 60)
        
        crear = input("¿Deseas crear esta copia de seguridad a doble cara automáticamente? (S/N): ").strip().upper()
        if crear in ["S", "SI"]:
            duplex_name = f"{default_printer} Doble Cara"
            log_message(f"Creando impresora duplicada '{duplex_name}'...")
            
            # Ejecutar PowerShell para crear la impresora y configurarla
            ps_cmd = f"""
            $source = Get-Printer -Name '{default_printer}'
            Add-Printer -Name '{duplex_name}' -DriverName $source.DriverName -PortName $source.PortName
            Set-PrintConfiguration -PrinterName '{duplex_name}' -DuplexingMode TwoSidedLongEdge
            """
            res = subprocess.run(["powershell", "-Command", ps_cmd], capture_output=True, text=True)
            if res.returncode == 0:
                log_message(f"--> Impresora '{duplex_name}' creada y configurada a DOBLE CARA con éxito.")
                
                # También forzar con win32print en la nueva impresora para estar seguros
                try:
                    PRINTER_ACCESS_USE = 0x00000008
                    handle = win32print.OpenPrinter(duplex_name, {"DesiredAccess": PRINTER_ACCESS_USE})
                    info = win32print.GetPrinter(handle, 9)
                    devmode = info['pDevMode']
                    if devmode is not None:
                        devmode.Duplex = 2
                        devmode.Fields = devmode.Fields | win32con.DM_DUPLEX
                        win32print.SetPrinter(handle, 9, info, 0)
                        log_message("--> Preferencias de usuario configuradas a doble cara vía win32print Level 9.")
                    win32print.ClosePrinter(handle)
                except Exception as e:
                    log_message(f"Advertencia al aplicar win32print Level 9 en la nueva impresora: {e}")
                    
                return duplex_name
            else:
                log_message(f"Error al crear la impresora duplicada: {res.stderr.strip()}")
        else:
            log_message("Creación de impresora omitida.")
    except Exception as e:
        log_message(f"Advertencia durante la creación de la impresora de doble cara: {e}")
        
    return None

def get_current_duplex_state(printer_name):
    """Obtiene la configuración original de duplex para poder restaurarla."""
    try:
        import win32print
        PRINTER_ACCESS_USE = 0x00000008
        try:
            handle = win32print.OpenPrinter(printer_name, {"DesiredAccess": PRINTER_ACCESS_USE})
        except Exception:
            handle = win32print.OpenPrinter(printer_name, {"DesiredAccess": win32print.PRINTER_ALL_ACCESS})
            
        # Intentar con Level 9
        try:
            info = win32print.GetPrinter(handle, 9)
            devmode = info['pDevMode']
            if devmode is not None:
                val = getattr(devmode, 'Duplex', None)
                if val is not None:
                    win32print.ClosePrinter(handle)
                    return "win32print_9", val
        except Exception:
            pass
            
        # Intentar con Level 2
        try:
            info = win32print.GetPrinter(handle, 2)
            devmode = info['pDevMode']
            if devmode is not None:
                val = getattr(devmode, 'Duplex', None)
                if val is not None:
                    win32print.ClosePrinter(handle)
                    return "win32print_2", val
        except Exception:
            pass
            
        win32print.ClosePrinter(handle)
    except Exception:
        pass

    try:
        cmd = f"(Get-PrintConfiguration -PrinterName '{printer_name}').DuplexingMode"
        res = subprocess.run(["powershell", "-Command", cmd], capture_output=True, text=True)
        if res.returncode == 0 and res.stdout.strip():
            return "powershell", res.stdout.strip()
    except Exception:
        pass
        
    return None, None

def set_duplex_state(printer_name, duplex=True):
    """Establece la impresión a doble cara (True) o normal (False) para la impresora."""
    log_message(f"Configurando dúplex ({duplex}) para '{printer_name}'...")
    
    # 1. Intentar con win32print usando Level 9 (Preferencias de usuario - No requiere Admin)
    try:
        import win32print
        import win32con
        
        PRINTER_ACCESS_USE = 0x00000008
        try:
            handle = win32print.OpenPrinter(printer_name, {"DesiredAccess": PRINTER_ACCESS_USE})
        except Exception:
            handle = win32print.OpenPrinter(printer_name, {"DesiredAccess": win32print.PRINTER_ALL_ACCESS})
            
        try:
            try:
                info = win32print.GetPrinter(handle, 9)
            except Exception:
                info2 = win32print.GetPrinter(handle, 2)
                info = {"pDevMode": info2["pDevMode"]}
                
            devmode = info['pDevMode']
            if devmode is not None:
                devmode.Duplex = 2 if duplex else 1
                devmode.Fields = devmode.Fields | win32con.DM_DUPLEX
                
                win32print.SetPrinter(handle, 9, info, 0)
                log_message("--> Dúplex configurado exitosamente vía win32print Level 9 (Preferencias de usuario).")
                win32print.ClosePrinter(handle)
                return True
        except Exception as e:
            log_message(f"Advertencia al configurar vía win32print Level 9: {e}")
            if 'handle' in locals():
                win32print.ClosePrinter(handle)
    except Exception as e:
        log_message(f"Advertencia general de win32print Level 9: {e}")

    # 2. Intentar con win32print usando Level 2 (Configuración global - Puede requerir Admin)
    try:
        import win32print
        import win32con
        handle = win32print.OpenPrinter(printer_name, {"DesiredAccess": win32print.PRINTER_ALL_ACCESS})
        try:
            info = win32print.GetPrinter(handle, 2)
            devmode = info['pDevMode']
            if devmode is not None:
                devmode.Duplex = 2 if duplex else 1
                devmode.Fields = devmode.Fields | win32con.DM_DUPLEX
                win32print.SetPrinter(handle, 2, info, 0)
                log_message("--> Dúplex configurado exitosamente vía win32print Level 2 (Configuración global).")
                win32print.ClosePrinter(handle)
                return True
        except Exception as e:
            log_message(f"Advertencia al configurar vía win32print Level 2: {e}")
            if 'handle' in locals():
                win32print.ClosePrinter(handle)
    except Exception as e:
        log_message(f"Advertencia general de win32print Level 2: {e}")

    # 3. Intentar con PowerShell como último recurso
    try:
        mode = "TwoSidedLongEdge" if duplex else "OneSided"
        cmd = f"Set-PrintConfiguration -PrinterName '{printer_name}' -DuplexingMode {mode}"
        res = subprocess.run(["powershell", "-Command", cmd], capture_output=True, text=True)
        if res.returncode == 0:
            log_message(f"--> Dúplex configurado exitosamente ({mode}) vía PowerShell.")
            return True
        else:
            log_message(f"Advertencia: PowerShell falló con código {res.returncode}: {res.stderr}")
    except Exception as e:
        log_message(f"Advertencia al ejecutar PowerShell: {e}")

    return False

def restore_duplex_state(printer_name, method, original_val):
    """Restaura la configuración de impresión a su estado original."""
    if not method or original_val is None:
        return
    log_message(f"Restaurando configuración original de la impresora '{printer_name}'...")
    
    if method == "win32print_9":
        try:
            import win32print
            import win32con
            PRINTER_ACCESS_USE = 0x00000008
            handle = win32print.OpenPrinter(printer_name, {"DesiredAccess": PRINTER_ACCESS_USE})
            try:
                info = win32print.GetPrinter(handle, 9)
            except Exception:
                info2 = win32print.GetPrinter(handle, 2)
                info = {"pDevMode": info2["pDevMode"]}
            devmode = info['pDevMode']
            if devmode is not None:
                devmode.Duplex = original_val
                devmode.Fields = devmode.Fields | win32con.DM_DUPLEX
                win32print.SetPrinter(handle, 9, info, 0)
                log_message("--> Configuración original restaurada vía win32print Level 9.")
            win32print.ClosePrinter(handle)
        except Exception as e:
            log_message(f"Error al restaurar vía Level 9: {e}")
            
    elif method == "win32print_2":
        try:
            import win32print
            import win32con
            handle = win32print.OpenPrinter(printer_name, {"DesiredAccess": win32print.PRINTER_ALL_ACCESS})
            info = win32print.GetPrinter(handle, 2)
            devmode = info['pDevMode']
            if devmode is not None:
                devmode.Duplex = original_val
                devmode.Fields = devmode.Fields | win32con.DM_DUPLEX
                win32print.SetPrinter(handle, 2, info, 0)
                log_message("--> Configuración original restaurada vía win32print Level 2.")
            win32print.ClosePrinter(handle)
        except Exception as e:
            log_message(f"Error al restaurar vía Level 2: {e}")
            
    elif method == "powershell":
        try:
            cmd = f"Set-PrintConfiguration -PrinterName '{printer_name}' -DuplexingMode {original_val}"
            subprocess.run(["powershell", "-Command", cmd], capture_output=True, text=True)
            log_message("--> Configuración original restaurada vía PowerShell.")
        except Exception as e:
            log_message(f"Error al restaurar vía PowerShell: {e}")

def process_car():
    print("=" * 60)
    print("  SUZUKI - CONFIGURADOR Y CONFIGURACIÓN RÁPIDA DE PDI")
    print("=" * 60)
    
    # 1. Resolver la ruta del archivo de Excel
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_file = os.path.join(script_dir, "FORMATO PDI.xlsx")
    
    if not os.path.exists(excel_file):
        excel_file = "FORMATO PDI.xlsx"
        if not os.path.exists(excel_file):
            log_user("Error: No se encontró el archivo 'FORMATO PDI.xlsx'.")
            return False

    # 2. Cargar el libro de trabajo (conservando fórmulas)
    log_user("Cargando formato Excel...")
    try:
        wb = openpyxl.load_workbook(excel_file)
    except Exception as e:
        log_user(f"Error al abrir el archivo Excel: {e}")
        return False

    sheet_names = wb.sheetnames

    # 3. Mostrar únicamente los modelos autorizados
    models_list = ["BALENO", "SWIFT", "DZIRE", "ERTIGA", "FRONX", "JIMNY"]
    
    print("\nModelos disponibles:")
    for idx, model_opt in enumerate(models_list, 1):
        print(f"  {idx}. {model_opt}")
        
    while True:
        try:
            sel = int(input(f"Selecciona el modelo (1-{len(models_list)}): "))
            if 1 <= sel <= len(models_list):
                selected_model = models_list[sel - 1]
                break
        except ValueError:
            pass
        print("Opción inválida. Intenta de nuevo.")

    # Pregunta especial para JIMNY: ¿3 o 5 puertas?
    jimny_doors = ""
    if selected_model == "JIMNY":
        print("\n¿Cuántas puertas tiene el Jimny?")
        print("  1. 3 PTAS")
        print("  2. 5 PTAS")
        while True:
            try:
                sel_doors = int(input("Selecciona (1-2): "))
                if sel_doors == 1:
                    jimny_doors = "3"
                    break
                elif sel_doors == 2:
                    jimny_doors = "5"
                    break
            except ValueError:
                pass
            print("Opción inválida. Intenta de nuevo.")

    # 4. Líneas predefinidas para cada modelo
    model_lines = {
        "BALENO": ["GLS", "GLX"],
        "SWIFT": ["GLS", "GLX", "SPORT"],
        "DZIRE": ["GLS", "GLX"],
        "ERTIGA": ["GLS", "GLX", "XL7"],
        "FRONX": ["GLS", "GLX"],
        "JIMNY": ["GLS", "GLX"]
    }
    
    lines_list = model_lines.get(selected_model, ["GLS", "GLX"])
    
    print(f"\nLíneas disponibles para {selected_model}:")
    for idx, line_opt in enumerate(lines_list, 1):
        print(f"  {idx}. {line_opt}")
        
    while True:
        try:
            sel = int(input(f"Selecciona la línea (1-{len(lines_list)}): "))
            if 1 <= sel <= len(lines_list):
                selected_line = lines_list[sel - 1]
                break
        except ValueError:
            pass
        print("Opción inválida. Intenta de nuevo.")

    # Preguntar por la transmisión (TA, TM o CVT)
    print("\nTransmisiones:")
    print("  1. TA")
    print("  2. TM")
    print("  3. CVT")
    while True:
        try:
            sel_trans = int(input("Selecciona la transmisión (1-3): "))
            if sel_trans == 1:
                selected_trans = "TA"
                break
            elif sel_trans == 2:
                selected_trans = "TM"
                break
            elif sel_trans == 3:
                selected_trans = "CVT"
                break
        except ValueError:
            pass
        print("Opción inválida. Intenta de nuevo.")

    # 5. Lógica de búsqueda de la hoja con fallback
    trans_search_order = [selected_trans]
    if selected_trans == "CVT":
        trans_search_order.append("TA")
    elif selected_trans == "TA":
        trans_search_order.append("CVT")

    matched_sheet_name = None
    
    # Buscar por coincidencia exacta de modelo + línea + transmisión (y filtro Jimny si aplica)
    for t_option in trans_search_order:
        for sheet_name in sheet_names:
            name_up = sheet_name.upper()
            if selected_model == "VITARA" and "GRAND VITARA" in name_up:
                continue
            if selected_model == "JIMNY" and not is_jimny_sheet_match(wb, sheet_name, jimny_doors):
                continue
            if (selected_model in name_up) and (selected_line in name_up) and (t_option in name_up):
                matched_sheet_name = sheet_name
                break
        if matched_sheet_name:
            break

    # Si no se encuentra, buscar solo por modelo + transmisión (ignorando la línea)
    if not matched_sheet_name:
        for t_option in trans_search_order:
            for sheet_name in sheet_names:
                name_up = sheet_name.upper()
                if selected_model == "VITARA" and "GRAND VITARA" in name_up:
                    continue
                if selected_model == "JIMNY" and not is_jimny_sheet_match(wb, sheet_name, jimny_doors):
                    continue
                if (selected_model in name_up) and (t_option in name_up):
                    matched_sheet_name = sheet_name
                    break
            if matched_sheet_name:
                break

    # Si aún no se encuentra, usar la primera hoja que coincida con el modelo (y filtro Jimny si aplica)
    if not matched_sheet_name:
        for sheet_name in sheet_names:
            name_up = sheet_name.upper()
            if selected_model == "VITARA" and "GRAND VITARA" in name_up:
                continue
            if selected_model == "JIMNY" and not is_jimny_sheet_match(wb, sheet_name, jimny_doors):
                continue
            if selected_model in name_up:
                matched_sheet_name = sheet_name
                break

    if not matched_sheet_name:
        log_user(f"Error: No se pudo encontrar ninguna plantilla para {selected_model}.")
        return False

    print(f"\n--> Plantilla seleccionada: '{matched_sheet_name}'")
    sheet = wb[matched_sheet_name]

    # Ajustar la página en openpyxl también por seguridad
    try:
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 2
    except Exception:
        pass

    # 6. Preguntar por los demás datos
    color = input("\n¿Qué color es? (ej. AZUL PASCAL, PLATA SIBERIA): ").strip().upper()
    key_num = input("¿Qué número de llave es? (opcional, presiona Enter para mantener la del formato): ").strip()

    # Obtener el VIN actual de la plantilla
    vin_template = sheet['D10'].value
    if not vin_template:
        vin_template = "MA3ZFEFS5VA376378" # Valor base fallback
    else:
        vin_template = str(vin_template).strip()

    # Preguntar por el dígito verificador después de la 'S'
    s_idx = vin_template.upper().find('S', 3)
    if s_idx != -1 and s_idx < len(vin_template) - 1:
        char_after_s_default = vin_template[s_idx + 1]
        print(f"\nVIN base: {vin_template}")
        print(f"La letra 'S' está en la posición {s_idx + 1}.")
        val_after_s = input(f"¿Qué número o letra tiene después de la S (actualmente '{char_after_s_default}')? ").strip().upper()
        if not val_after_s:
            val_after_s = char_after_s_default
        check_digit_idx = s_idx + 1
    else:
        char_after_s_default = vin_template[8] if len(vin_template) > 8 else "X"
        print(f"\nVIN base: {vin_template}")
        val_after_s = input(f"¿Qué número o letra tiene en la 9ª posición del VIN (actualmente '{char_after_s_default}')? ").strip().upper()
        if not val_after_s:
            val_after_s = char_after_s_default
        check_digit_idx = 8

    # Preguntar por la terminación del VIN
    while True:
        vin_end = input("\nIntroduce los últimos números del VIN (terminación, ej. 375215): ").strip()
        if vin_end.isdigit():
            break
        print("La terminación debe contener solo números. Intenta de nuevo.")

    # 7. Construir el nuevo VIN
    vin_len = len(vin_template)
    suffix_len = len(vin_end)
    
    part1 = vin_template[:check_digit_idx]
    part2 = val_after_s
    part3 = vin_template[check_digit_idx + 1 : vin_len - suffix_len]
    part4 = vin_end
    
    new_vin = part1 + part2 + part3 + part4
    log_user(f"--> Nuevo VIN generado: {new_vin}")

    # 8. Modificar la hoja de Excel
    clean_sheet_title = matched_sheet_name.replace("OK", "").replace("(2)", "").strip()
    sheet['D9'].value = f"{clean_sheet_title} {color}".strip()
    sheet['D10'].value = new_vin
    
    if key_num:
        sheet['D11'].value = key_num

    # Establecer la hoja modificada como la pestaña activa
    wb.active = wb.sheetnames.index(matched_sheet_name)

    # 9. Hacer respaldo y guardar archivo
    backup_file = os.path.join(script_dir, "FORMATO PDI_backup.xlsx")
    log_message(f"Creando respaldo en: {backup_file}")
    try:
        shutil.copyfile(excel_file, backup_file)
    except Exception as e:
        log_message(f"Advertencia: No se pudo crear el archivo de respaldo: {e}")

    log_user("Guardando cambios en Excel...")
    try:
        wb.save(excel_file)
        wb.close()
        log_user("¡Cambios guardados con éxito!")
    except Exception as e:
        log_user(f"Error al guardar los cambios: {e}")
        log_user("Asegúrate de que el archivo Excel no esté abierto en otro programa.")
        return False

    # 9.5 LIMPIAR CACHÉ DE IMPRESORA DEL ZIP (.xlsx)
    strip_printer_settings_from_zip(excel_file)

    # 10. Comando de impresión rápida para Windows
    log_user("\n" + "=" * 40)
    log_user("              IMPRESIÓN")
    log_user("=" * 40)
    
    abs_excel_path = os.path.abspath(excel_file)
    
    try:
        import win32com.client
        import win32print
        
        # 10a. Configurar la impresora para impresión a doble cara
        # Buscar si hay un perfil duplicado de doble cara o crearlo si lo autoriza el usuario
        duplex_printer = ensure_duplex_printer_profile()
        
        if duplex_printer:
            printer_name = duplex_printer
            method, original_val = None, None
            log_user(f"Uso de impresora de doble cara seleccionada: '{printer_name}'")
        else:
            printer_name = win32print.GetDefaultPrinter()
            log_user(f"Impresora predeterminada detectada: '{printer_name}'")
            method, original_val = get_current_duplex_state(printer_name)
            
            # Cambiar el duplex en el driver
            set_duplex_state(printer_name, duplex=True)

        # 10b. Imprimir la hoja usando Excel COM
        log_message("Iniciando instancia aislada de Excel en segundo plano...")
        excel_app = win32com.client.DispatchEx("Excel.Application")
        excel_app.Visible = False
        
        try:
            workbook = excel_app.Workbooks.Open(abs_excel_path)
            
            # Forzar ActivePrinter después de abrir el libro para obligar a Excel a recargar el driver
            try:
                set_excel_active_printer(excel_app, printer_name)
            except Exception as pe:
                log_message(f"Advertencia: No se pudo asignar ActivePrinter en Excel: {pe}")

            if workbook is not None:
                active_sheet = workbook.ActiveSheet
                if active_sheet is not None:
                    # Configurar para que la hoja activa se ajuste exactamente a 1 página de ancho por 2 de alto
                    try:
                        active_sheet.PageSetup.Zoom = False
                        active_sheet.PageSetup.FitToPagesWide = 1
                        active_sheet.PageSetup.FitToPagesTall = 2
                        log_message("Ajuste de página configurado: Ancho = 1 pág., Alto = 2 págs. (1 hoja física doble cara).")
                    except Exception as pe:
                        log_message(f"Advertencia: No se pudo configurar el ajuste de tamaño de página: {pe}")
                    
                    log_user(f"Mandando a imprimir la hoja activa: '{active_sheet.Name}'...")
                    active_sheet.PrintOut()
                else:
                    log_user("Error: No se pudo acceder a la hoja activa.")
                workbook.Close(False)
                log_user("¡Impresión enviada correctamente!")
            else:
                log_user("Error: Excel no pudo abrir el libro de trabajo.")
        except Exception as e:
            log_user(f"Error durante el proceso de impresión con Excel: {e}")
        finally:
            excel_app.Quit()

        # 10c. Restaurar la configuración original de duplex de la impresora
        if method and original_val is not None:
            restore_duplex_state(printer_name, method, original_val)
            
    except ImportError:
        log_user("Los módulos 'pywin32' o 'win32print' no están instalados en este sistema Python.")
        if hasattr(os, 'startfile'):
            log_user("Intentando enviar la impresión usando el comando estándar de Windows...")
            try:
                os.startfile(abs_excel_path, "print")
                log_user("¡Comando de impresión de Windows ejecutado!")
            except Exception as e:
                log_user(f"No se pudo imprimir automáticamente: {e}")
        else:
            log_user("El comando de impresión rápida de Windows (os.startfile) no está disponible en este sistema operativo.")
            
    return True

def main():
    # Elevar privilegios al inicio en Windows para garantizar modificaciones en win32print y SetPrinter
    check_and_elevate()

    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        log_file = os.path.join(script_dir, "pdi_debug_log.txt")
        if os.path.exists(log_file):
            os.remove(log_file)
    except Exception:
        pass

    while True:
        process_car()
        
        print("\n" + "-" * 60)
        opcion = input("¿Deseas configurar otro automóvil? (S/N): ").strip().upper()
        if opcion not in ["S", "SI"]:
            print("\n¡Gracias por utilizar la herramienta Suzuki PDI Configurator!")
            print("Proceso finalizado.")
            input("Presiona Enter para salir...")
            break
        print("\n" * 2)

if __name__ == "__main__":
    main()
